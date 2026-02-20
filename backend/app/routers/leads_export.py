"""
Lead Export API Endpoints
Supports CSV and Excel export with field selection and filtering
ASYNC VERSION with DEBUG LOGGING
"""
from fastapi import APIRouter, Depends, Query, HTTPException
from fastapi.responses import StreamingResponse
from sqlalchemy.ext.asyncio import AsyncSession
from sqlalchemy import select
from typing import List, Optional
import csv
import io
from datetime import datetime
import logging
import traceback

# Configure logging
logger = logging.getLogger(__name__)
logger.setLevel(logging.DEBUG)

# For Excel support
try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment
    EXCEL_AVAILABLE = True
    logger.info("✅ Excel export available (openpyxl installed)")
except ImportError:
    EXCEL_AVAILABLE = False
    logger.warning("⚠️ Excel export NOT available (openpyxl not installed)")

from app.database import get_db
from app.rbac import require_view_leads
from app.models import User
from app.models import Lead

router = APIRouter()


# Available fields for export
EXPORTABLE_FIELDS = {
    'email': 'Email',
    'first_name': 'First Name',
    'last_name': 'Last Name',
    'full_name': 'Full Name',
    'phone': 'Phone',
    'job_title': 'Job Title',
    'linkedin_url': 'LinkedIn URL',
    'company_name': 'Company Name',
    'company_website': 'Company Website',
    'company_industry': 'Industry',
    'status': 'Status',
    'source_name': 'Source',
    'fit_score': 'Fit Score',
    'email_verified': 'Email Verified',
    'email_deliverability_score': 'Deliverability Score',
    'review_decision': 'Review Decision',
    'review_notes': 'Review Notes',
    'reviewed_by': 'Reviewed By',
    'reviewed_at': 'Reviewed At',
    'created_at': 'Created At',
    'updated_at': 'Updated At',
}


def get_field_value(lead: Lead, field: str) -> str:
    """Extract and format field value from lead"""
    try:
        value = None
        
        if field == 'full_name':
            if lead.first_name and lead.last_name:
                value = f"{lead.first_name} {lead.last_name}"
            elif lead.first_name:
                value = lead.first_name
            elif lead.last_name:
                value = lead.last_name
        elif field == 'fit_score':
            value = f"{lead.fit_score * 100:.1f}%" if lead.fit_score else None
        elif field == 'email_deliverability_score':
            value = f"{lead.email_deliverability_score * 100:.1f}%" if lead.email_deliverability_score else None
        elif field == 'email_verified':
            value = "Yes" if lead.email_verified else "No"
        elif field in ['created_at', 'updated_at', 'reviewed_at']:
            dt = getattr(lead, field, None)
            value = dt.strftime('%Y-%m-%d %H:%M:%S') if dt else None
        else:
            value = getattr(lead, field, None)
        
        return str(value) if value is not None else ''
    except Exception as e:
        logger.error(f"Error getting field '{field}' value: {e}")
        return ''


def generate_csv(leads: List[Lead], fields: List[str]) -> io.StringIO:
    """Generate CSV file from leads"""
    try:
        logger.info(f"Generating CSV with {len(leads)} leads and {len(fields)} fields")
        output = io.StringIO()
        writer = csv.writer(output, quoting=csv.QUOTE_MINIMAL)
        
        # Write header
        headers = [EXPORTABLE_FIELDS.get(field, field) for field in fields]
        writer.writerow(headers)
        logger.debug(f"CSV headers: {headers}")
        
        # Write data rows
        for idx, lead in enumerate(leads):
            row = [get_field_value(lead, field) for field in fields]
            writer.writerow(row)
            if idx == 0:
                logger.debug(f"First row sample: {row[:3]}...")
        
        output.seek(0)
        logger.info(f"✅ CSV generated successfully")
        return output
    except Exception as e:
        logger.error(f"❌ Error generating CSV: {e}")
        logger.error(traceback.format_exc())
        raise


def generate_excel(leads: List[Lead], fields: List[str]) -> io.BytesIO:
    """Generate Excel file from leads with formatting"""
    try:
        logger.info(f"Generating Excel with {len(leads)} leads and {len(fields)} fields")
        
        if not EXCEL_AVAILABLE:
            raise HTTPException(
                status_code=500,
                detail="Excel export not available. Install openpyxl: pip install openpyxl"
            )
        
        wb = Workbook()
        ws = wb.active
        ws.title = "Leads"
        
        # Header styling
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF", size=11)
        header_alignment = Alignment(horizontal="left", vertical="center")
        
        # Write headers
        headers = [EXPORTABLE_FIELDS.get(field, field) for field in fields]
        for col_idx, header in enumerate(headers, start=1):
            cell = ws.cell(row=1, column=col_idx, value=header)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = header_alignment
        logger.debug(f"Excel headers: {headers}")
        
        # Write data rows
        for row_idx, lead in enumerate(leads, start=2):
            for col_idx, field in enumerate(fields, start=1):
                value = get_field_value(lead, field)
                ws.cell(row=row_idx, column=col_idx, value=value)
            if row_idx == 2:
                logger.debug(f"First row written")
        
        # Auto-adjust column widths
        for col_idx, field in enumerate(fields, start=1):
            max_length = len(EXPORTABLE_FIELDS.get(field, field))
            for row_idx in range(2, min(len(leads) + 2, 100)):  # Check first 100 rows
                cell_value = str(ws.cell(row=row_idx, column=col_idx).value or '')
                max_length = max(max_length, len(cell_value))
            
            # Set column width (max 50 characters)
            ws.column_dimensions[ws.cell(row=1, column=col_idx).column_letter].width = min(max_length + 2, 50)
        
        # Freeze header row
        ws.freeze_panes = "A2"
        
        # Save to BytesIO
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        logger.info(f"✅ Excel generated successfully")
        return output
    except Exception as e:
        logger.error(f"❌ Error generating Excel: {e}")
        logger.error(traceback.format_exc())
        raise


@router.get("/export")
async def export_leads(
    format: str = Query("csv", regex="^(csv|xlsx)$"),
    fields: Optional[str] = Query(None, description="Comma-separated list of fields to export"),
    # Filter parameters
    status: Optional[str] = None,
    source: Optional[str] = None,
    email_verified: Optional[bool] = None,
    search: Optional[str] = None,
    sort_by: str = "created_at",
    sort_order: str = "desc",
    limit: Optional[int] = None,
    # Auth
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(require_view_leads),
):
    """
    Export leads to CSV or Excel format
    
    - **format**: Export format (csv or xlsx)
    - **fields**: Comma-separated list of fields (default: all fields)
    - **Filters**: All standard lead filters apply
    - **limit**: Maximum number of leads to export (default: no limit)
    """
    try:
        logger.info(f"📥 Export request: format={format}, user={current_user.email}")
        logger.debug(f"Filters: status={status}, source={source}, email_verified={email_verified}, search={search}")
        
        # Parse fields
        if fields:
            selected_fields = [f.strip() for f in fields.split(',') if f.strip() in EXPORTABLE_FIELDS]
            logger.info(f"Selected fields: {selected_fields}")
        else:
            # Default fields for export
            selected_fields = [
                'email', 'first_name', 'last_name', 'phone', 'job_title',
                'company_name', 'company_website', 'company_industry',
                'status', 'source_name', 'fit_score', 'email_verified',
                'created_at'
            ]
            logger.info(f"Using default fields: {len(selected_fields)} fields")
        
        if not selected_fields:
            logger.warning("No valid fields selected")
            raise HTTPException(
                status_code=400,
                detail="No valid fields selected for export"
            )
        
        # Build query using select() for AsyncSession
        logger.info("Building query...")
        query = select(Lead)
        
        # Add tenant filtering if model supports it
        if hasattr(Lead, 'tenant_id') and hasattr(current_user, 'tenant_id'):
            query = query.where(Lead.tenant_id == current_user.tenant_id)
            logger.debug(f"Added tenant filter: {current_user.tenant_id}")
        else:
            logger.debug("No tenant filtering (tenant_id not found)")
        
        # Apply filters
        if status:
            query = query.where(Lead.status == status)
            logger.debug(f"Filter: status={status}")
        if source:
            query = query.where(Lead.source_name == source)
            logger.debug(f"Filter: source={source}")
        if email_verified is not None:
            query = query.where(Lead.email_verified == email_verified)
            logger.debug(f"Filter: email_verified={email_verified}")
        if search:
            search_filter = f"%{search}%"
            query = query.where(
                (Lead.email.ilike(search_filter)) |
                (Lead.first_name.ilike(search_filter)) |
                (Lead.last_name.ilike(search_filter)) |
                (Lead.company_name.ilike(search_filter))
            )
            logger.debug(f"Filter: search={search}")
        
        # Apply sorting
        try:
            sort_field = getattr(Lead, sort_by)
            if sort_order == "desc":
                query = query.order_by(sort_field.desc())
            else:
                query = query.order_by(sort_field.asc())
            logger.debug(f"Sorting: {sort_by} {sort_order}")
        except AttributeError:
            logger.warning(f"Sort field '{sort_by}' not found, using default")
            query = query.order_by(Lead.created_at.desc())
        
        # Apply limit
        if limit:
            query = query.limit(limit)
            logger.debug(f"Limit: {limit}")
        
        # Execute query and fetch leads
        logger.info("Executing query...")
        result = await db.execute(query)
        leads = result.scalars().all()
        logger.info(f"✅ Found {len(leads)} leads")
        
        if not leads:
            logger.warning("No leads found")
            raise HTTPException(
                status_code=404,
                detail="No leads found matching the criteria"
            )
        
        # Generate file based on format
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        
        if format == "xlsx":
            logger.info("Generating Excel file...")
            file_content = generate_excel(leads, selected_fields)
            filename = f"leads_export_{timestamp}.xlsx"
            media_type = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        else:  # csv
            logger.info("Generating CSV file...")
            file_content = generate_csv(leads, selected_fields)
            filename = f"leads_export_{timestamp}.csv"
            media_type = "text/csv"
        
        logger.info(f"✅ Export complete: {filename}")
        
        # Return file as streaming response
        return StreamingResponse(
            file_content,
            media_type=media_type,
            headers={
                "Content-Disposition": f"attachment; filename={filename}",
                "X-Total-Leads": str(len(leads)),
            }
        )
        
    except HTTPException:
        # Re-raise HTTP exceptions
        raise
    except Exception as e:
        # Log and raise unexpected errors
        logger.error(f"❌ EXPORT FAILED: {e}")
        logger.error(f"Error type: {type(e).__name__}")
        logger.error(traceback.format_exc())
        raise HTTPException(
            status_code=500,
            detail=f"Export failed: {str(e)}"
        )


@router.get("/export/fields")
async def get_exportable_fields(
    current_user: User = Depends(require_view_leads),
):
    """
    Get list of available fields for export
    Returns field keys and display names
    """
    return {
        "fields": [
            {"key": key, "label": label}
            for key, label in EXPORTABLE_FIELDS.items()
        ]
    }